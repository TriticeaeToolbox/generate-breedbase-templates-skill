"""
Breedbase Template Generator

Generates three Breedbase upload templates from diverse input data formats:
1. Accessions template
2. Trials template
3. Observations template

This module handles various input formats including multi-sheet Excel, multi-file Excel,
pre-structured CSV, simple Excel, and multi-rep Excel formats.
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any
import re
import warnings
warnings.filterwarnings('ignore')


class BreedbaseTemplateGenerator:
    """Main class for generating Breedbase upload templates"""

    def __init__(self, input_folder: str, assets_folder: str = None,
                 program: str = 'PROGRAM', experiment_code: str = None, year: int = None):
        """
        Initialize the generator

        Args:
            input_folder: Path to folder containing input data files
            assets_folder: Path to assets folder with ontologies and definitions
            program: Breeding program name (used in breeding_program column)
            experiment_code: Code for trial/plot naming (defaults to program)
            year: Trial year (defaults to current year)
        """
        self.input_folder = Path(input_folder)
        self.assets_folder = Path(assets_folder) if assets_folder else Path(__file__).parent / 'assets'
        self.program = program
        self.experiment_code = experiment_code if experiment_code else program
        self.year = year if year else datetime.now().year

        # Load reference data
        self.ontologies = self._load_ontologies()
        self.trait_abbreviations = self._load_trait_abbreviations()
        self.accession_columns = self._load_column_definitions('accessions')
        self.trial_columns = self._load_column_definitions('trials')

        # Storage for parsed data
        self.all_accessions = []
        self.all_trials = []
        self.all_observations = []

    def _load_ontologies(self) -> Dict[str, Dict]:
        """Load trait ontologies from OBO files"""
        ontologies = {}
        ontology_dir = self.assets_folder / 'trait-ontologies'

        for species_file in ['oat.txt', 'barley.txt', 'wheat.txt']:
            species = species_file.replace('.txt', '')
            file_path = ontology_dir / species_file

            if file_path.exists():
                ontologies[species] = self._parse_obo_file(file_path)

        return ontologies

    def _parse_obo_file(self, file_path: Path) -> Dict:
        """Parse OBO format ontology file"""
        ontology = {'traits': {}, 'variables': {}}

        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Split into terms
        terms = re.split(r'\n\[Term\]', content)

        for term_block in terms[1:]:  # Skip header
            term_data = {}

            for line in term_block.split('\n'):
                line = line.strip()
                if ':' in line:
                    key, value = line.split(':', 1)
                    key = key.strip()
                    value = value.strip()

                    if key == 'id':
                        term_data['id'] = value
                    elif key == 'name':
                        term_data['name'] = value
                    elif key == 'def':
                        term_data['def'] = value.strip('"').split('"')[0]
                    elif key == 'synonym':
                        if 'synonyms' not in term_data:
                            term_data['synonyms'] = []
                        # Extract synonym from quoted text
                        syn_match = re.search(r'"([^"]+)"', value)
                        if syn_match:
                            term_data['synonyms'].append(syn_match.group(1))

            if 'id' in term_data and 'name' in term_data:
                term_id = term_data['id']
                term_name = term_data['name'].lower()

                # Store both by ID and by name for lookup
                ontology['traits'][term_name] = term_data
                ontology['variables'][term_id] = term_data

        return ontology

    def _load_trait_abbreviations(self) -> pd.DataFrame:
        """Load trait abbreviations mapping"""
        abbr_file = self.assets_folder / 'trait_abbreviations.xlsx'

        if abbr_file.exists():
            df = pd.read_excel(abbr_file)
            return df
        return pd.DataFrame()

    def _load_column_definitions(self, template_type: str) -> pd.DataFrame:
        """Load column definitions for templates"""
        col_def_dir = self.assets_folder / 'upload-template-column-definitions'

        if template_type == 'accessions':
            file_path = col_def_dir / 'accessions-template-column-definitions.csv'
        elif template_type == 'trials':
            file_path = col_def_dir / 'traits-template-column-definitions.csv'
        else:
            return pd.DataFrame()

        if file_path.exists():
            return pd.read_csv(file_path)
        return pd.DataFrame()

    def _detect_format(self, file_path: Path) -> str:
        """
        Detect input file format

        Returns: format type ('multi_sheet', 'simple_excel', 'csv', 'multi_rep', etc.)
        """
        if file_path.suffix.lower() == '.csv':
            # Check if pre-structured
            df = pd.read_csv(file_path, nrows=5)
            if 'trial_name' in df.columns and 'accession_name' in df.columns:
                return 'pre_structured_csv'
            return 'csv'

        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            # Multi-sheet with location names
            if len(sheet_names) > 3 and any('overall' in s.lower() for s in sheet_names):
                return 'multi_sheet_locations'

            # Multi-file format (has Cover and Page 2)
            if 'Cover' in sheet_names and 'Page 2' in sheet_names:
                return 'cover_page2'

            # Check for replicate structure
            df = pd.read_excel(file_path, nrows=10)
            if any('rep' in str(col).lower() for col in df.columns):
                return 'multi_rep'

            return 'simple_excel'

        return 'unknown'

    def _find_header_row(self, df: pd.DataFrame, keywords: List[str] = None) -> int:
        """
        Find the header row in a DataFrame by looking for common keywords

        Args:
            df: DataFrame to search
            keywords: List of keywords to look for (e.g., ['entry', 'variety', 'name'])

        Returns: Row index of header, or 0 if not found
        """
        if keywords is None:
            keywords = ['entry', 'variety', 'name', 'accession', 'plot', 'trial']

        for idx, row in df.iterrows():
            row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
            if any(kw in row_str for kw in keywords):
                return int(idx)
        return 0

    def _detect_species(self, data_source: Any) -> str:
        """
        Detect species from data source (file content, metadata, etc.)

        Returns: Species name (e.g., 'Avena sativa', 'Triticum aestivum', 'Hordeum vulgare')
        """
        # Search for species indicators in text
        text = str(data_source).lower()

        if 'oat' in text:
            return 'Avena sativa'
        elif 'wheat' in text:
            return 'Triticum aestivum'
        elif 'barley' in text:
            return 'Hordeum vulgare'

        # Default to oats if not detected
        return 'Avena sativa'

    def _extract_pedigree_components(self, pedigree: str) -> Tuple[Optional[str], Optional[str]]:
        """
        Extract female and male parent from pedigree string

        Args:
            pedigree: Pedigree string (e.g., 'FEMALE/MALE' or 'FEMALE//MALE/PARENT')

        Returns: (female_parent, male_parent) tuple
        """
        if pd.isna(pedigree) or not pedigree:
            return None, None

        pedigree = str(pedigree).strip()

        # Simple biparental cross: FEMALE/MALE
        if '/' in pedigree:
            parts = pedigree.split('/')
            if len(parts) >= 2:
                female = parts[0].strip()
                male = parts[1].strip()
                return female if female else None, male if male else None

        return None, None

    def _match_trait_to_ontology(self, trait_name: str, unit: str, species: str = 'oat') -> Optional[Dict]:
        """
        Match a trait name and unit to an ontology term

        Args:
            trait_name: Name of trait (may be abbreviated)
            unit: Unit of measurement
            species: Species name

        Returns: Dict with ontology_id, full_name, and description
        """
        if species not in self.ontologies:
            return None

        ontology = self.ontologies[species]
        trait_name_clean = trait_name.lower().strip()

        # Try exact match first
        if trait_name_clean in ontology['traits']:
            trait_data = ontology['traits'][trait_name_clean]
            return {
                'ontology_id': trait_data['id'],
                'full_name': trait_data['name'],
                'description': f"{trait_data['name']} - {unit}",
                'unit': unit
            }

        # Try abbreviation lookup
        if not self.trait_abbreviations.empty:
            abbr_matches = self.trait_abbreviations[
                self.trait_abbreviations.iloc[:, 0].str.lower() == trait_name_clean
            ]
            if not abbr_matches.empty:
                full_name = abbr_matches.iloc[0, 1]
                if full_name.lower() in ontology['traits']:
                    trait_data = ontology['traits'][full_name.lower()]
                    return {
                        'ontology_id': trait_data['id'],
                        'full_name': trait_data['name'],
                        'description': f"{trait_data['name']} - {unit}",
                        'unit': unit
                    }

        # Try fuzzy matching with synonyms
        for trait_key, trait_data in ontology['traits'].items():
            # Check synonyms
            if 'synonyms' in trait_data:
                for syn in trait_data['synonyms']:
                    if syn.lower() == trait_name_clean:
                        return {
                            'ontology_id': trait_data['id'],
                            'full_name': trait_data['name'],
                            'description': f"{trait_data['name']} - {unit}",
                            'unit': unit
                        }

            # Check if trait name contains the key
            if trait_key in trait_name_clean or trait_name_clean in trait_key:
                return {
                    'ontology_id': trait_data['id'],
                    'full_name': trait_data['name'],
                    'description': f"{trait_data['name']} - {unit}",
                    'unit': unit
                }

        return None

    def _convert_units(self, value: Any, from_unit: str, to_unit: str, species: str = 'oat') -> float:
        """
        Convert measurement units

        Args:
            value: Original value
            from_unit: Source unit
            to_unit: Target unit
            species: Species (for yield conversion factors)

        Returns: Converted value
        """
        if pd.isna(value):
            return value

        try:
            value = float(value)
        except:
            return value

        from_unit = from_unit.lower().strip()
        to_unit = to_unit.lower().strip()

        # Yield conversions
        if 'bu/a' in from_unit and 'g/m' in to_unit:
            # bu/A to g/m² (oat-specific factor)
            return value * 3.586715
        elif 'kg/ha' in from_unit and 'g/m' in to_unit:
            # kg/ha to g/m²
            return value / 10.0

        # Test weight conversions
        elif 'lb/bu' in from_unit and 'g/l' in to_unit:
            # lb/bu to g/L
            return value * 12.871981

        # Height conversions
        elif ('inch' in from_unit or 'in' == from_unit) and 'cm' in to_unit:
            # inches to cm
            return value * 2.54

        # Date to Julian day
        elif 'date' in from_unit.lower() and 'julian' in to_unit.lower():
            # Convert date to Julian day
            if isinstance(value, (datetime, pd.Timestamp)):
                return value.timetuple().tm_yday
            return value

        # No conversion needed
        return value

    def _date_to_julian(self, date_value: Any) -> Optional[int]:
        """Convert date to Julian day (1-365/366)"""
        if pd.isna(date_value):
            return None

        if isinstance(date_value, (datetime, pd.Timestamp)):
            return date_value.timetuple().tm_yday

        # Try parsing string
        if isinstance(date_value, str):
            try:
                dt = pd.to_datetime(date_value)
                return dt.timetuple().tm_yday
            except:
                pass

        # If already a number, assume it's Julian day
        try:
            return int(float(date_value))
        except:
            return None

    def generate_all_templates(self, output_folder: str = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Main method to generate all three templates

        Args:
            output_folder: Optional folder to save output files

        Returns: (accessions_df, trials_df, observations_df)
        """
        # Process all input files
        input_files = []
        for ext in ['*.xlsx', '*.xls', '*.csv']:
            input_files.extend(self.input_folder.glob(ext))

        for file_path in input_files:
            print(f"Processing {file_path.name}...")
            format_type = self._detect_format(file_path)
            print(f"  Detected format: {format_type}")

            if format_type == 'multi_sheet_locations':
                self._process_multi_sheet_locations(file_path)
            elif format_type == 'cover_page2':
                self._process_cover_page2(file_path)
            elif format_type == 'pre_structured_csv':
                self._process_pre_structured_csv(file_path)
            elif format_type == 'simple_excel':
                self._process_simple_excel(file_path)
            elif format_type == 'multi_rep':
                self._process_multi_rep(file_path)
            elif format_type == 'csv':
                self._process_csv(file_path)

        # Generate final templates
        accessions_df = generate_accessions_template(self.all_accessions)
        trials_df = generate_trials_template(self.all_trials)
        observations_df = generate_observations_template(self.all_observations)

        # Save if output folder specified
        if output_folder:
            output_path = Path(output_folder)
            output_path.mkdir(parents=True, exist_ok=True)

            accessions_df.to_excel(output_path / 'accessions.xls', index=False)
            trials_df.to_excel(output_path / 'trials.xls', index=False)
            observations_df.to_excel(output_path / 'observations.xls', index=False)

            print(f"\nTemplates saved to {output_folder}/")

        return accessions_df, trials_df, observations_df

    def _process_multi_sheet_locations(self, file_path: Path):
        """Process Format 1: Multi-sheet Excel with location sheets"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Identify location sheets (not "Overall_" sheets)
        location_sheets = [s for s in wb.sheetnames if not s.lower().startswith('overall')]

        # Extract pedigrees from Overall sheets if available
        pedigrees = {}
        for sheet_name in wb.sheetnames:
            if 'overall' in sheet_name.lower():
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Look for pedigree columns
                for col in df.columns:
                    if 'pedigree' in str(col).lower():
                        for idx, row in df.iterrows():
                            if 'name' in df.columns:
                                acc_name = row.get('name') or row.get('Name')
                                if acc_name and pd.notna(row[col]):
                                    pedigrees[acc_name] = row[col]

        # Process each location sheet
        for sheet_name in location_sheets:
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # Extract metadata from top rows
            trial_description = ""
            if len(df) > 0:
                trial_description = str(df.iloc[0, 0]) if pd.notna(df.iloc[0, 0]) else ""

            # Find header row
            header_idx = self._find_header_row(df)

            # Re-read with proper headers
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)

            # Detect species
            species = self._detect_species(trial_description)

            # Parse data rows
            self._extract_accessions_trials_observations(
                df,
                location=sheet_name,
                trial_description=trial_description,
                species=species,
                pedigrees=pedigrees
            )

        wb.close()

    def _process_cover_page2(self, file_path: Path):
        """Process Format 2: Cover + Page 2 sheets"""
        # Extract location from filename
        location_match = re.search(r'_([a-z_]+)\.xlsx?$', file_path.name, re.I)
        location = location_match.group(1).title() if location_match else "Unknown"

        # Read metadata from Cover sheet
        cover_df = pd.read_excel(file_path, sheet_name='Cover')
        metadata = {}
        for idx, row in cover_df.iterrows():
            if len(row) >= 2:
                key = str(row.iloc[0]).strip()
                value = str(row.iloc[1]).strip()
                if key and value and key != 'nan':
                    metadata[key] = value

        # Read data from Page 2
        df = pd.read_excel(file_path, sheet_name='Page 2')

        # Find header row
        header_idx = self._find_header_row(df)
        df = pd.read_excel(file_path, sheet_name='Page 2', header=header_idx)

        # Detect species
        species = self._detect_species(' '.join(metadata.values()))

        self._extract_accessions_trials_observations(
            df,
            location=location,
            trial_description=metadata.get('Trial', 'Trial'),
            species=species,
            metadata=metadata
        )

    def _process_pre_structured_csv(self, file_path: Path):
        """Process Format 3: Pre-structured CSV"""
        df = pd.read_csv(file_path)

        # Split into trial columns and trait columns
        trial_cols = [c for c in df.columns if c in ['trial_name', 'breeding_program', 'location',
                                                       'year', 'design_type', 'accession_name',
                                                       'plot_number', 'block_number', 'rep_number',
                                                       'plot_name', 'entry_number', 'row_number',
                                                       'col_number']]
        trait_cols = [c for c in df.columns if c not in trial_cols]

        # Extract species
        species = self._detect_species(df.get('trial_name', [''])[0] if 'trial_name' in df.columns else '')

        # Process each row
        for idx, row in df.iterrows():
            # Extract accession
            if 'accession_name' in row:
                accession = {
                    'accession_name': row['accession_name'],
                    'species_name': species
                }
                self.all_accessions.append(accession)

            # Extract trial data
            trial_data = {}
            for col in trial_cols:
                if col in row:
                    trial_data[col] = row[col]
            if trial_data:
                self.all_trials.append(trial_data)

            # Extract observation data
            obs_data = {'observationunit_name': row.get('plot_name', '')}
            for col in trait_cols:
                if pd.notna(row[col]):
                    obs_data[col] = row[col]
            if obs_data:
                self.all_observations.append(obs_data)

    def _process_simple_excel(self, file_path: Path):
        """Process Format 4: Simple single-sheet Excel"""
        # Read first few rows to detect structure
        df_preview = pd.read_excel(file_path, nrows=10)

        # Find header and unit rows
        header_idx = self._find_header_row(df_preview)

        # Read full data
        df = pd.read_excel(file_path, header=header_idx)

        # Check if next row contains units
        df_check = pd.read_excel(file_path, header=None)
        if header_idx + 1 < len(df_check):
            units_row = df_check.iloc[header_idx + 1]
            # Re-read skipping unit row
            df = pd.read_excel(file_path, header=header_idx, skiprows=[header_idx + 1])

        species = self._detect_species(file_path.name)

        self._extract_accessions_trials_observations(
            df,
            location="Unknown",
            trial_description=f"Trial from {file_path.name}",
            species=species
        )

    def _process_multi_rep(self, file_path: Path):
        """Process Format 5: Multi-replicate Excel with complex headers"""
        # Read to detect structure
        df_preview = pd.read_excel(file_path, nrows=10, header=None)

        # Find where actual data starts
        data_start_row = 0
        for idx in range(len(df_preview)):
            row_str = ' '.join(str(v) for v in df_preview.iloc[idx] if pd.notna(v)).lower()
            if 'entry' in row_str or 'variety' in row_str:
                data_start_row = idx
                break

        # Read with proper header
        df = pd.read_excel(file_path, header=data_start_row)

        species = self._detect_species(file_path.name)

        self._extract_accessions_trials_observations(
            df,
            location="Unknown",
            trial_description=f"Multi-rep trial from {file_path.name}",
            species=species,
            has_reps=True
        )

    def _process_csv(self, file_path: Path):
        """Process generic CSV file"""
        df = pd.read_csv(file_path)

        # Find header row if needed
        header_idx = self._find_header_row(df)
        if header_idx > 0:
            df = pd.read_csv(file_path, header=header_idx)

        species = self._detect_species(file_path.name)

        self._extract_accessions_trials_observations(
            df,
            location=file_path.stem,
            trial_description=f"Trial from {file_path.name}",
            species=species
        )

    def _extract_accessions_trials_observations(self, df: pd.DataFrame, location: str,
                                                 trial_description: str, species: str,
                                                 pedigrees: Dict = None, metadata: Dict = None,
                                                 has_reps: bool = False):
        """
        Extract accessions, trials, and observations from a DataFrame

        This is the core extraction logic that works across formats
        """
        if pedigrees is None:
            pedigrees = {}
        if metadata is None:
            metadata = {}

        # Identify key columns dynamically
        name_col = None
        entry_col = None
        variety_col = None

        for col in df.columns:
            col_lower = str(col).lower()
            if 'name' in col_lower and not name_col:
                name_col = col
            elif 'entry' in col_lower and not entry_col:
                entry_col = col
            elif 'variety' in col_lower or 'selection' in col_lower:
                variety_col = col

        # Use variety > name > entry as accession name source
        accession_col = variety_col or name_col or entry_col

        if not accession_col:
            print(f"  Warning: Could not find accession name column in {location}")
            return

        # Generate trial name using experiment_code for naming, program for column
        year = metadata.get('year', self.year)
        program = metadata.get('breeding_program', self.program)
        trial_name = f"{self.experiment_code}_{year}_{location.replace(' ', '')}"

        # Identify trait columns (exclude structural columns)
        structural_keywords = ['entry', 'name', 'variety', 'selection', 'pedigree',
                               'number', 'origin', 'years']
        trait_columns = []
        for col in df.columns:
            col_lower = str(col).lower()
            if not any(kw in col_lower for kw in structural_keywords):
                trait_columns.append(col)

        # Process each row
        plot_number = 101
        for idx, row in df.iterrows():
            accession_name = row.get(accession_col)

            if pd.isna(accession_name) or not str(accession_name).strip():
                continue

            accession_name = str(accession_name).strip()

            # Extract accession
            pedigree = pedigrees.get(accession_name, row.get('pedigree', None))
            female, male = self._extract_pedigree_components(pedigree)

            accession_data = {
                'accession_name': accession_name,
                'species_name': species
            }
            if pedigree:
                accession_data['purdy pedigree'] = pedigree
            if female:
                accession_data['female_parent'] = female
            if male:
                accession_data['male_parent'] = male

            self.all_accessions.append(accession_data)

            # Extract trial data
            plot_name = f"{trial_name}-PLOT_{plot_number}"
            trial_data = {
                'trial_name': trial_name,
                'breeding_program': program,
                'location': location,
                'year': year,
                'design_type': metadata.get('design_type', 'RCBD'),
                'description': trial_description,
                'accession_name': accession_name,
                'plot_number': plot_number,
                'block_number': metadata.get('block_number', 1),
                'plot_name': plot_name
            }

            if entry_col and pd.notna(row.get(entry_col)):
                trial_data['entry_number'] = row[entry_col]

            self.all_trials.append(trial_data)

            # Extract observations
            obs_data = {'observationunit_name': plot_name}

            for trait_col in trait_columns:
                value = row.get(trait_col)
                if pd.notna(value):
                    # Try to match to ontology
                    # For now, store with original column name
                    # Will be mapped in generate_observations_template
                    obs_data[trait_col] = value

            self.all_observations.append(obs_data)
            plot_number += 1


def generate_accessions_template(accessions_data: List[Dict]) -> pd.DataFrame:
    """
    Generate accessions template from extracted data

    Args:
        accessions_data: List of accession dictionaries

    Returns: DataFrame with accessions template
    """
    if not accessions_data:
        return pd.DataFrame(columns=['accession_name', 'species_name'])

    # Convert to DataFrame and remove duplicates
    df = pd.DataFrame(accessions_data)

    # Remove duplicates based on accession_name, keeping first occurrence
    df = df.drop_duplicates(subset=['accession_name'], keep='first')

    # Ensure required columns are present
    if 'accession_name' not in df.columns:
        df['accession_name'] = ''
    if 'species_name' not in df.columns:
        df['species_name'] = 'Avena sativa'

    # Order columns: required first, then optional
    required_cols = ['accession_name', 'species_name']
    optional_cols = [c for c in df.columns if c not in required_cols]

    final_columns = required_cols + optional_cols
    df = df[final_columns]

    return df


def generate_trials_template(trials_data: List[Dict]) -> pd.DataFrame:
    """
    Generate trials template from extracted data

    Args:
        trials_data: List of trial dictionaries

    Returns: DataFrame with trials template
    """
    if not trials_data:
        return pd.DataFrame()

    df = pd.DataFrame(trials_data)

    # Ensure required columns
    required_cols = ['trial_name', 'breeding_program', 'location', 'year',
                     'design_type', 'description', 'accession_name',
                     'plot_number', 'block_number', 'plot_name']

    for col in required_cols:
        if col not in df.columns:
            df[col] = ''

    # Order columns
    all_cols = required_cols + [c for c in df.columns if c not in required_cols]
    df = df[all_cols]

    return df


def generate_observations_template(observations_data: List[Dict]) -> pd.DataFrame:
    """
    Generate observations template from extracted data

    Args:
        observations_data: List of observation dictionaries

    Returns: DataFrame with observations template (with ontology IDs in column headers)
    """
    if not observations_data:
        return pd.DataFrame()

    df = pd.DataFrame(observations_data)

    # Ensure observationunit_name is first column
    if 'observationunit_name' in df.columns:
        cols = ['observationunit_name'] + [c for c in df.columns if c != 'observationunit_name']
        df = df[cols]

    # TODO: Map trait columns to ontology IDs
    # This would require matching column names to ontology terms
    # For now, return with original column names
    # Full implementation would rename columns to format:
    # "{trait_name} - {description}|{ONTOLOGY_ID}"

    return df


def generate_templates_from_folder(input_folder: str, output_folder: str = None,
                                   assets_folder: str = None,
                                   program: str = 'PROGRAM',
                                   experiment_code: str = None,
                                   year: int = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Convenience function to generate all templates from a folder

    Args:
        input_folder: Path to folder containing input data files
        output_folder: Optional path to save output templates
        assets_folder: Optional path to assets folder (defaults to ./assets)
        program: Breeding program name (used in breeding_program column)
        experiment_code: Code for trial/plot naming (defaults to program)
        year: Trial year (defaults to current year)

    Returns: (accessions_df, trials_df, observations_df)
    """
    generator = BreedbaseTemplateGenerator(input_folder, assets_folder, program, experiment_code, year)
    return generator.generate_all_templates(output_folder)


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python generate_breedbase_templates.py <input_folder> [output_folder]")
        sys.exit(1)

    input_folder = sys.argv[1]
    output_folder = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"Generating Breedbase templates from: {input_folder}")
    accessions, trials, observations = generate_templates_from_folder(input_folder, output_folder)

    print(f"\nGenerated templates:")
    print(f"  Accessions: {len(accessions)} rows")
    print(f"  Trials: {len(trials)} rows")
    print(f"  Observations: {len(observations)} rows")
